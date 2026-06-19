"""Excel→構造化 JSON 変換の決定論検証(.xlsx は openpyxl, .xls は xlrd)。"""

from __future__ import annotations

import io
import json

import openpyxl
import pytest

from app.converters import convert


def _xlsx_bytes(sheets: dict[str, list[list]]) -> bytes:
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    for name, rows in sheets.items():
        worksheet = workbook.create_sheet(title=name)
        for row in rows:
            worksheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _payload(source: bytes) -> dict:
    outcome = convert(source, "", "excel_to_json", None)
    assert outcome.converted is True
    assert outcome.derived_bytes is not None
    return json.loads(outcome.derived_bytes.decode("utf-8"))


def test_xlsx_single_sheet_records() -> None:
    payload = _payload(_xlsx_bytes({"Sheet1": [["name", "age"], ["Alice", 30], ["Bob", 25]]}))
    assert payload["sheet_count"] == 1
    sheet = payload["sheets"][0]
    assert sheet["name"] == "Sheet1"
    assert sheet["columns"] == ["name", "age"]
    assert sheet["row_count"] == 2
    # 整数 float は小数点を落とす。
    assert sheet["rows"][0] == {"name": "Alice", "age": "30"}


def test_xlsx_multiple_sheets_preserved_in_order() -> None:
    payload = _payload(
        _xlsx_bytes({"A": [["x"], ["1"]], "B": [["y"], ["2"]]})
    )
    assert [sheet["name"] for sheet in payload["sheets"]] == ["A", "B"]


def test_xlsx_is_deterministic() -> None:
    source = _xlsx_bytes({"S": [["a", "b"], [1, 2]]})
    first = convert(source, "", "excel_to_json", None).derived_bytes
    second = convert(source, "", "excel_to_json", None).derived_bytes
    assert first == second


def test_duplicate_headers_and_short_rows() -> None:
    payload = _payload(_xlsx_bytes({"S": [["id", "id", "note"], [1, 2]]}))
    sheet = payload["sheets"][0]
    assert sheet["columns"] == ["id", "id_1", "note"]
    assert sheet["rows"][0] == {"id": "1", "id_1": "2", "note": ""}


def test_non_excel_profile_passes_through() -> None:
    outcome = convert(_xlsx_bytes({"S": [["a"], ["1"]]}), "", "csv_to_json", None)
    assert outcome.converted is False


def test_empty_input_passes_through() -> None:
    outcome = convert(b"", "", "excel_to_json", None)
    assert outcome.converted is False
    assert "excel_empty" in outcome.warnings


def test_xls_via_xlrd() -> None:
    xlwt = pytest.importorskip("xlwt")  # .xls 書き込みで往復検証(無ければ skip)
    workbook = xlwt.Workbook()
    worksheet = workbook.add_sheet("Sheet1")
    for r, row in enumerate([["name", "score"], ["Carol", 99]]):
        for c, value in enumerate(row):
            worksheet.write(r, c, value)
    buffer = io.BytesIO()
    workbook.save(buffer)
    payload = _payload(buffer.getvalue())
    sheet = payload["sheets"][0]
    assert sheet["columns"] == ["name", "score"]
    assert sheet["rows"][0] == {"name": "Carol", "score": "99"}
